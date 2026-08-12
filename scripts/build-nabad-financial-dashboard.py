#!/usr/bin/env python3
"""
Build NabadAi founder financial dashboard (.xlsx).
Run: python3 scripts/build-nabad-financial-dashboard.py
Output: docs/NabadAi-Financial-Dashboard.xlsx
"""
from __future__ import annotations

import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "NabadAi-Financial-Dashboard.xlsx"

# --- NabadAi brand colors ---
C_BG = "0B0F17"
C_CARD = "121826"
C_INPUT = "1A2234"
C_PURPLE = "7C5CFF"
C_TEAL = "23D5AB"
C_TEXT = "E8EEF7"
C_MUTED = "8B95A8"
C_GREEN = "22C55E"
C_RED = "EF4444"
C_YELLOW = "F59E0B"
C_WHITE = "FFFFFF"

fill_bg = PatternFill("solid", fgColor=C_BG)
fill_card = PatternFill("solid", fgColor=C_CARD)
fill_input = PatternFill("solid", fgColor=C_INPUT)
fill_purple = PatternFill("solid", fgColor=C_PURPLE)
fill_teal = PatternFill("solid", fgColor=C_TEAL)
fill_green = PatternFill("solid", fgColor="14532D")
fill_red = PatternFill("solid", fgColor="450A0A")
fill_yellow = PatternFill("solid", fgColor="451A03")
fill_header = PatternFill("solid", fgColor="1E293B")

font_title = Font(name="Calibri", size=20, bold=True, color=C_WHITE)
font_h1 = Font(name="Calibri", size=14, bold=True, color=C_TEAL)
font_h2 = Font(name="Calibri", size=11, bold=True, color=C_PURPLE)
font_label = Font(name="Calibri", size=10, color=C_MUTED)
font_value = Font(name="Calibri", size=11, color=C_TEXT)
font_kpi = Font(name="Calibri", size=22, bold=True, color=C_WHITE)
font_kpi_sm = Font(name="Calibri", size=14, bold=True, color=C_TEXT)
font_input = Font(name="Calibri", size=11, color=C_WHITE)
font_note = Font(name="Calibri", size=9, italic=True, color=C_MUTED)

thin = Side(style="thin", color="334155")
border_card = Border(left=thin, right=thin, top=thin, bottom=thin)
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_right = Alignment(horizontal="right", vertical="center")


def style_range(ws, cell_range, fill=None, font=None, border=None, alignment=None):
    for row in ws[cell_range]:
        for cell in row:
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            if border:
                cell.border = border
            if alignment:
                cell.alignment = alignment


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def mark_input(cell, fmt=None):
    cell.fill = fill_input
    cell.font = font_input
    cell.border = border_card
    if fmt:
        cell.number_format = fmt


def mark_calc(cell, fmt=None):
    cell.fill = fill_card
    cell.font = font_value
    cell.border = border_card
    if fmt:
        cell.number_format = fmt


def section_title(ws, row, col, text, span=4):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)
    c = ws.cell(row=row, column=col, value=text)
    c.font = font_h1
    c.fill = fill_bg
    c.alignment = align_left


def kpi_block(ws, row, col, label, formula, fmt='#,##0.00'):
    ws.cell(row=row, column=col, value=label).font = font_label
    ws.cell(row=row, column=col).fill = fill_bg
    c = ws.cell(row=row + 1, column=col, value=formula)
    c.font = font_kpi_sm
    c.fill = fill_card
    c.border = border_card
    c.number_format = fmt
    c.alignment = align_center
    return c


def build_assumptions(wb):
    ws = wb.create_sheet("ASSUMPTIONS", 0)
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [36, 18, 14, 42])

    ws["A1"] = "NABAD AI — ASSUMPTIONS (edit yellow cells)"
    ws["A1"].font = font_title
    ws["A1"].fill = fill_bg
    ws.merge_cells("A1:D1")

    rows = [
        ("", "Value", "Unit", "Notes / Source"),
        ("— PRICING —", "", "", ""),
        ("Weekly subscription price", 3.99, "USD", "pro-plan-config.js"),
        ("Monthly subscription price", 9.99, "USD", "pro-plan-config.js"),
        ("Weekly credits granted", 400, "credits", "400/week incl. trial week"),
        ("Monthly credits granted", 1200, "credits", "1000 + 200 bonus"),
        ("Credits per full song", 12, "credits", "2 variants A+B per debit"),
        ("", "", "", ""),
        ("— APPLE APP STORE —", "", "", ""),
        ("Apple commission year 1", 0.30, "%", "Default 30%"),
        ("Apple commission after year 1", 0.15, "%", "After 12 mo continuous sub"),
        ("Apple Small Business Program (15% from start)?", 0, "0=No 1=Yes", "Edit to 1 if enrolled"),
        ("Share of subs on Apple (current mix)", 0.80, "%", "Editable"),
        ("", "", "", ""),
        ("— STRIPE UAE —", "", "", ""),
        ("Stripe % fee", 0.029, "%", "2.9% domestic default"),
        ("Stripe fixed fee (AED)", 1, "AED", "Per successful charge"),
        ("Stripe Billing fee", 0.007, "%", "0.7% billing volume"),
        ("Intl card surcharge", 0, "%", "Optional"),
        ("Currency conversion surcharge", 0, "%", "Optional"),
        ("AED / USD exchange rate", 3.67, "AED per USD", "Editable"),
        ("Share of subs on Stripe/Web", 0.20, "%", "Should sum to 100% with Apple"),
        ("", "", "", ""),
        ("— AI / API COSTS (from codebase) —", "", "", ""),
        ("Suno USD per API credit", 0.00525, "USD", "$5.25 / 1000 credits — music-generation-log.js"),
        ("Suno cost per generation (12 cr)", 0.063, "USD", "docs/nabadai-music-cost-model.csv"),
        ("Lyria cost per track (admin/spike)", 0.08, "USD", "Not default production path"),
        ("Gemini Coach cost per message", 0.003, "USD", "Editable estimate — coach.js"),
        ("Gemini Photo Mood per analysis", 0.002, "USD", "Editable — image-mood.js"),
        ("Pollinations cover art", 0, "USD", "Free tier default"),
        ("Avg Coach messages / subscriber / mo", 25, "messages", "Editable"),
        ("Avg Photo Mood uses / subscriber / mo", 2, "uses", "Editable"),
        ("Credit utilization % (subs use credits)", 0.65, "%", "Not all credits spent"),
        ("Heavy user credit utilization", 0.95, "%", "Worst-case subscriber"),
        ("Remix/mix of features vs pure songs", 0.85, "%", "Weight toward 12-cr song cost"),
        ("Historical Suno credit pack spend (total)", 5.25, "USD", "Edit your actual purchases"),
        ("Historical other API spend (total)", 0, "USD", "Gemini etc. to date"),
        ("", "", "", ""),
        ("— SUBSCRIBER BEHAVIOR —", "", "", ""),
        ("Current paid users (total)", 0, "users", "Edit live count"),
        ("Current Weekly subscribers", 0, "users", ""),
        ("Current Monthly subscribers", 0, "users", ""),
        ("Monthly churn rate", 0.08, "%", "Editable"),
        ("Avg customer lifetime (months)", 6, "months", "Or use 1/churn"),
        ("Marketing spend this month", 0, "USD", ""),
        ("", "", "", ""),
        ("— FIXED COSTS SUMMARY (also in Expenses) —", "", "", ""),
        ("Default infra start months ago", 3, "months", "~3 mo dev subscriptions"),
    ]

    start_row = 3
    for i, row in enumerate(rows):
        r = start_row + i
        for j, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=j, value=val)
            cell.fill = fill_bg
        if row[0] and str(row[0]).startswith("—"):
            ws.cell(row=r, column=1).font = font_h2
        elif row[0] and row[0] not in ("", "Value"):
            ws.cell(row=r, column=1).font = font_label
        if row[1] != "" and row[0] and not str(row[0]).startswith("—") and row[0] != "Value":
            mark_input(ws.cell(row=r, column=2))
        ws.cell(row=r, column=3).font = font_note
        ws.cell(row=r, column=4).font = font_note

    # Map assumption labels to rows for named references
    label_to_row = {}
    for i, row in enumerate(rows):
        if row[0] and not str(row[0]).startswith("—") and row[0] not in ("", "Value"):
            label_to_row[row[0]] = start_row + i

    ws.freeze_panes = "A3"
    return ws, label_to_row


def a_row(label_to_row, label):
    return label_to_row[label]


def embed(ref: str) -> str:
    """Use inside a formula when ref may already start with '='."""
    s = str(ref).strip()
    return s[1:] if s.startswith("=") else s


def as_formula(ref: str) -> str:
    """Ensure a cell reference/expression is a full formula."""
    s = str(ref).strip()
    return s if s.startswith("=") else f"={s}"


def build_expenses(wb, label_to_row):
    ws = wb.create_sheet("EXPENSE TRACKER")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [28, 16, 16, 12, 14, 10, 12, 16, 16, 28])

    headers = [
        "Expense", "Category", "Provider", "Start Date", "Monthly Cost",
        "Currency", "Months Paid", "Total Paid To Date", "Current Monthly", "Notes",
    ]
    ws.append(["NABAD AI — EXPENSE TRACKER"])
    ws["A1"].font = font_title
    ws.merge_cells("A1:J1")
    ws.append(headers)
    for c in range(1, 11):
        cell = ws.cell(row=2, column=c)
        cell.font = font_h2
        cell.fill = fill_header
        cell.alignment = align_center

    # Pre-fill from user prompt — rows 3+
    data = [
        ("Vercel hosting", "Recurring Fixed", "Vercel", "2025-11-01", 20, "USD", 3, "=E3*G3", "=IF(G3>0,E3,0)", "Preview + production"),
        ("Supabase", "Recurring Fixed", "Supabase", "2025-11-01", 25, "USD", 3, "=E4*G4", "=IF(G4>0,E4,0)", "DB + auth + storage"),
        ("Cursor IDE", "Recurring Fixed", "Cursor", "2025-11-01", 70, "USD", 3, "=E5*G5", "=IF(G5>0,E5,0)", "Dev subscription"),
        ("Shopify", "Recurring Fixed", "Shopify", "2025-11-01", 15, "USD", 3, "=E6*G6", "=IF(G6>0,E6,0)", ""),
        ("Domain (nabadai.com)", "Recurring Fixed", "Registrar", "2025-11-01", "=10/12", "USD", 3, "=E7*G7", "=IF(G7>0,E7,0)", "$10/year amortized"),
        ("Cursor one-time purchase", "One-Time/Startup", "Cursor", "", 0, "USD", 1, 200, 0, "$200 one-time — not monthly"),
        ("Apple Developer Program", "One-Time/Startup", "Apple", "2025-11-01", 0, "USD", 1, 99, 0, "Paid once; renewal separate"),
        ("Dubai trade license", "One-Time/Startup", "UAE", "2025-11-01", 0, "AED", 1, "=1070/'ASSUMPTIONS'!$B$36", 0, "AED 1,070 one-time"),
        ("Marketing (this month)", "Marketing", "Various", "", f"='ASSUMPTIONS'!B{label_to_row['Marketing spend this month']}", "USD", 1, "=E11", "=E11", "Link to assumptions or edit"),
        ("Historical API / Suno credits", "Variable/API", "Suno", "", 0, "USD", 1, f"='ASSUMPTIONS'!B{label_to_row['Historical Suno credit pack spend (total)']}", 0, "From assumptions"),
        ("Historical other API", "Variable/API", "Gemini/etc", "", 0, "USD", 1, f"='ASSUMPTIONS'!B{label_to_row['Historical other API spend (total)']}", 0, ""),
    ]

    r0 = 3
    for i, row in enumerate(data):
        r = r0 + i
        for j, val in enumerate(row, 1):
            ws.cell(row=r, column=j, value=val)
        # Input styling for editable columns
        for col in (1, 2, 3, 4, 5, 6, 7, 10):
            if col == 8 and isinstance(row[7], str) and row[7].startswith("="):
                mark_calc(ws.cell(row=r, column=8), "$#,##0.00")
            elif col in (5, 7) or (col == 4 and row[3]):
                mark_input(ws.cell(row=r, column=col), "$#,##0.00" if col == 5 else "0")
            elif col == 4:
                mark_input(ws.cell(row=r, column=col), "YYYY-MM-DD")
        mark_calc(ws.cell(row=r, column=8), "$#,##0.00")
        mark_calc(ws.cell(row=r, column=9), "$#,##0.00")

    # Totals row
    tr = r0 + len(data) + 1
    ws.cell(row=tr, column=1, value="TOTALS").font = font_h2
    ws.cell(row=tr, column=8, value=f"=SUM(H{r0}:H{tr-1})")
    ws.cell(row=tr, column=9, value=f"=SUM(I{r0}:I{tr-1})")
    mark_calc(ws.cell(row=tr, column=8), "$#,##0.00")
    mark_calc(ws.cell(row=tr, column=9), "$#,##0.00")

    ws.freeze_panes = "A3"
    return ws, r0, tr


def build_investment(wb, exp_r0, exp_tr):
    ws = wb.create_sheet("TOTAL INVESTMENT")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [40, 18, 40])

    ws["A1"] = "TOTAL NABAD AI INVESTMENT TO DATE"
    ws["A1"].font = font_title
    ws.merge_cells("A1:C1")

    categories = [
        ("Total Infrastructure (recurring paid to date)", f"=SUMIF('EXPENSE TRACKER'!B:B,\"Recurring Fixed\", 'EXPENSE TRACKER'!H:H)"),
        ("Total Development (Cursor recurring + one-time)", f"=SUMIF('EXPENSE TRACKER'!B:B,\"Recurring Fixed\", 'EXPENSE TRACKER'!H:H)+SUMIF('EXPENSE TRACKER'!A:A,\"Cursor one-time*\", 'EXPENSE TRACKER'!H:H)"),
        ("Total AI/API Investment (historical)", f"=SUMIF('EXPENSE TRACKER'!B:B,\"Variable/API\", 'EXPENSE TRACKER'!H:H)"),
        ("Total Legal/Setup (Apple + trade license)", f"=SUMIF('EXPENSE TRACKER'!B:B,\"One-Time/Startup\", 'EXPENSE TRACKER'!H:H)-SUMIF('EXPENSE TRACKER'!A:A,\"Cursor one-time*\", 'EXPENSE TRACKER'!H:H)"),
        ("Total Marketing Investment (historical)", f"=SUMIF('EXPENSE TRACKER'!B:B,\"Marketing\", 'EXPENSE TRACKER'!H:H)"),
        ("", ""),
        ("TOTAL NABAD AI INVESTMENT TO DATE", f"=SUM('EXPENSE TRACKER'!H{exp_r0}:'EXPENSE TRACKER'!H{exp_tr-1})"),
    ]

    row = 3
    for label, formula in categories:
        ws.cell(row=row, column=1, value=label).font = font_h2 if "TOTAL NABAD" in label else font_label
        ws.cell(row=row, column=1).fill = fill_bg
        if formula:
            c = ws.cell(row=row, column=2, value=formula)
            mark_calc(c, "$#,##0.00")
            if "TOTAL NABAD" in label:
                c.font = font_kpi
        row += 1

    ws["A12"] = "Note: Historical = cash already spent. Current monthly burn is on Founder Dashboard."
    ws["A12"].font = font_note
    return ws


def build_ai_unit_economics(wb, lr):
    ws = wb.create_sheet("AI UNIT ECONOMICS")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [22, 22, 14, 12, 16, 14, 16, 14, 18, 18])

    ws["A1"] = "AI UNIT ECONOMICS — from NabadAi codebase"
    ws["A1"].font = font_title
    ws.merge_cells("A1:J1")

    headers = [
        "Provider", "Feature", "User credits", "Nabad $/unit", "Notes",
        "Avg gen/user/mo", "Avg $/user/mo", "Heavy user $/mo",
        "Weekly sub est. AI $", "Monthly sub est. AI $",
    ]
    ws.append([""])
    ws.append(headers)
    for c in range(1, 11):
        ws.cell(row=3, column=c).font = font_h2
        ws.cell(row=3, column=c).fill = fill_header

    ar = lr
    W_CRED = f"ASSUMPTIONS!B{ar['Weekly credits granted']}"
    M_CRED = f"ASSUMPTIONS!B{ar['Monthly credits granted']}"
    SUNO_CR = f"ASSUMPTIONS!B{ar['Suno USD per API credit']}"
    UTIL = f"ASSUMPTIONS!B{ar['Credit utilization % (subs use credits)']}"
    HEAVY = f"ASSUMPTIONS!B{ar['Heavy user credit utilization']}"
    COACH = f"ASSUMPTIONS!B{ar['Gemini Coach cost per message']}"
    COACH_N = f"ASSUMPTIONS!B{ar['Avg Coach messages / subscriber / mo']}"
    PHOTO = f"ASSUMPTIONS!B{ar['Gemini Photo Mood per analysis']}"
    PHOTO_N = f"ASSUMPTIONS!B{ar['Avg Photo Mood uses / subscriber / mo']}"
    GEN12 = f"ASSUMPTIONS!B{ar['Suno cost per generation (12 cr)']}"

    features = [
        ("Suno (default)", "Full song (2 variants)", 12, f"={GEN12}", "api/suno/generate.js"),
        ("Suno", "Remix / hum track", 10, f"=10*{SUNO_CR}", "stems add_instrumental"),
        ("Suno", "Mashup", 12, f"={GEN12}", "music/mashup.js"),
        ("Suno", "Persona save", 5, f"=5*{SUNO_CR}", "persona.js"),
        ("Suno", "Instrumental / stems", 2, f"=2*{SUNO_CR}", "stems.js"),
        ("Suno", "Sound loop", 2.5, f"=2.5*{SUNO_CR}", "sounds.js"),
        ("Gemini", "Coach chat", 0, f"={COACH}", "Free to user; Pro unlimited"),
        ("Gemini", "Photo Mood analysis", 0, f"={PHOTO}", "Free before song debit"),
        ("Pollinations", "Cover art", 0, 0, "Default free"),
        ("Lyria (spike)", "Full song alt", 12, f"=ASSUMPTIONS!B{ar['Lyria cost per track (admin/spike)']}", "Admin only — not default"),
    ]

    r = 4
    for prov, feat, cred, cost, note in features:
        ws.cell(row=r, column=1, value=prov)
        ws.cell(row=r, column=2, value=feat)
        ws.cell(row=r, column=3, value=cred)
        ws.cell(row=r, column=4, value=cost)
        ws.cell(row=r, column=5, value=note)
        mark_calc(ws.cell(row=r, column=4), "$0.0000")
        r += 1

    # Summary block
    sr = r + 2
    section_title(ws, sr, 1, "SUBSCRIBER AI COST ESTIMATES (weighted by credit utilization)", 6)
    sr += 1

    ws.cell(row=sr, column=1, value="Blended Suno cost per Nabad credit").font = font_label
    ws.cell(row=sr, column=2, value=f"={SUNO_CR}")
    mark_calc(ws.cell(row=sr, column=2), "$0.00000")
    sr += 1

    ws.cell(row=sr, column=1, value="Weekly — avg AI cost (Suno credits + Gemini)").font = font_label
    ws.cell(row=sr, column=2, value=f"={W_CRED}*{UTIL}*{SUNO_CR}+{COACH_N}*{COACH}+{PHOTO_N}*{PHOTO}")
    mark_calc(ws.cell(row=sr, column=2), "$0.00")
    ws.cell(row=sr, column=3, value="At assumed utilization")
    sr += 1

    ws.cell(row=sr, column=1, value="Monthly — avg AI cost").font = font_label
    ws.cell(row=sr, column=2, value=f"={M_CRED}*{UTIL}*{SUNO_CR}+{COACH_N}*{COACH}+{PHOTO_N}*{PHOTO}")
    mark_calc(ws.cell(row=sr, column=2), "$0.00")
    sr += 1

    ws.cell(row=sr, column=1, value="Weekly — heavy user AI cost").font = font_label
    ws.cell(row=sr, column=2, value=f"={W_CRED}*{HEAVY}*{SUNO_CR}+{COACH_N}*2*{COACH}+{PHOTO_N}*2*{PHOTO}")
    mark_calc(ws.cell(row=sr, column=2), "$0.00")
    sr += 1

    ws.cell(row=sr, column=1, value="Monthly — heavy user AI cost").font = font_label
    ws.cell(row=sr, column=2, value=f"={M_CRED}*{HEAVY}*{SUNO_CR}+{COACH_N}*2*{COACH}+{PHOTO_N}*2*{PHOTO}")
    mark_calc(ws.cell(row=sr, column=2), "$0.00")

    # Store summary row refs for dashboard
    ws.meta_avg_weekly_row = sr - 3
    ws.meta_avg_monthly_row = sr - 2
    ws.meta_heavy_weekly_row = sr - 1
    ws.meta_heavy_monthly_row = sr
    return ws


def build_subscription_economics(wb, lr, ai_ws):
    ws = wb.create_sheet("SUBSCRIPTION ECONOMICS")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [24, 14, 14, 14, 14, 14, 14, 14, 14, 14])

    ws["A1"] = "SUBSCRIPTION ECONOMICS — Apple vs Stripe"
    ws["A1"].font = font_title
    ws.merge_cells("A1:J1")

    P_W = f"ASSUMPTIONS!B{lr['Weekly subscription price']}"
    P_M = f"ASSUMPTIONS!B{lr['Monthly subscription price']}"
    A1 = f"ASSUMPTIONS!B{lr['Apple commission year 1']}"
    A2 = f"ASSUMPTIONS!B{lr['Apple commission after year 1']}"
    SMB = f"ASSUMPTIONS!B{lr['Apple Small Business Program (15% from start)?']}"
    apple_comm_expr = f"IF({SMB}=1,{A2},{A1})"
    ST_P = f"ASSUMPTIONS!B{lr['Stripe % fee']}"
    ST_F = f"ASSUMPTIONS!B{lr['Stripe fixed fee (AED)']}"
    ST_B = f"ASSUMPTIONS!B{lr['Stripe Billing fee']}"
    FX = f"ASSUMPTIONS!B{lr['AED / USD exchange rate']}"
    AI_W = f"'AI UNIT ECONOMICS'!B{ai_ws.meta_avg_weekly_row}"
    AI_M = f"'AI UNIT ECONOMICS'!B{ai_ws.meta_avg_monthly_row}"

    stripe_fee_expr = lambda price: f"={price}*{ST_P}+({ST_F}/{FX})+{price}*{ST_B}"

    headers = ["Plan / Channel", "Gross $", "Platform fee $", "Net revenue $", "Est. AI cost $", "Contribution $", "Margin %", "Songs/mo incl.", "Credits/mo", "Notes"]
    ws.append([""])
    ws.append(headers)
    for c in range(1, 11):
        ws.cell(row=3, column=c).font = font_h2
        ws.cell(row=3, column=c).fill = fill_header

    # Row 4-7 plan economics
    ws.cell(row=4, column=1, value="Weekly — Apple")
    ws.cell(row=4, column=2, value=f"={P_W}")
    ws.cell(row=4, column=3, value=f"={P_W}*({apple_comm_expr})")
    ws.cell(row=4, column=4, value="=B4-C4")
    ws.cell(row=4, column=5, value=f"={AI_W}")
    ws.cell(row=4, column=6, value="=D4-E4")
    ws.cell(row=4, column=7, value="=IF(D4>0,F4/D4,0)")
    ws.cell(row=4, column=8, value=f"=ASSUMPTIONS!B{lr['Weekly credits granted']}/ASSUMPTIONS!B{lr['Credits per full song']}")
    ws.cell(row=4, column=9, value=f"=ASSUMPTIONS!B{lr['Weekly credits granted']}")
    ws.cell(row=4, column=10, value="Year-1 Apple default")

    ws.cell(row=5, column=1, value="Monthly — Apple")
    ws.cell(row=5, column=2, value=f"={P_M}")
    ws.cell(row=5, column=3, value=f"={P_M}*({apple_comm_expr})")
    ws.cell(row=5, column=4, value="=B5-C5")
    ws.cell(row=5, column=5, value=f"={AI_M}")
    ws.cell(row=5, column=6, value="=D5-E5")
    ws.cell(row=5, column=7, value="=IF(D5>0,F5/D5,0)")

    ws.cell(row=6, column=1, value="Weekly — Stripe")
    ws.cell(row=6, column=2, value=f"={P_W}")
    ws.cell(row=6, column=3, value=stripe_fee_expr(P_W))
    ws.cell(row=6, column=4, value="=B6-C6")
    ws.cell(row=6, column=5, value=f"={AI_W}")
    ws.cell(row=6, column=6, value="=D6-E6")
    ws.cell(row=6, column=7, value="=IF(D6>0,F6/D6,0)")

    ws.cell(row=7, column=1, value="Monthly — Stripe")
    ws.cell(row=7, column=2, value=f"={P_M}")
    ws.cell(row=7, column=3, value=stripe_fee_expr(P_M))
    ws.cell(row=7, column=4, value="=B7-C7")
    ws.cell(row=7, column=5, value=f"={AI_M}")
    ws.cell(row=7, column=6, value="=D7-E7")
    ws.cell(row=7, column=7, value="=IF(D7>0,F7/D7,0)")

    for r in range(4, 8):
        for col in range(2, 8):
            mark_calc(ws.cell(row=r, column=col), "$0.00" if col < 7 else "0.0%")

    # Profitability flags
    pr = 9
    section_title(ws, pr, 1, "PROFITABILITY CHECK (contribution > 0?)", 6)
    pr += 1
    for label, ref in [
        ("Weekly $3.99 profitable on Apple?", "=IF('SUBSCRIPTION ECONOMICS'!F4>0,\"YES\",\"NO — raise price or cut usage\")"),
        ("Monthly $9.99 profitable on Apple?", "=IF('SUBSCRIPTION ECONOMICS'!F5>0,\"YES\",\"NO\")"),
        ("Weekly profitable on Stripe?", "=IF('SUBSCRIPTION ECONOMICS'!F6>0,\"YES\",\"NO\")"),
        ("Monthly profitable on Stripe?", "=IF('SUBSCRIPTION ECONOMICS'!F7>0,\"YES\",\"NO\")"),
        ("Best channel for Weekly", "=IF(F6>F4,\"Stripe\",\"Apple\")"),
        ("Best channel for Monthly", "=IF(F7>F5,\"Stripe\",\"Apple\")"),
    ]:
        ws.cell(row=pr, column=1, value=label).font = font_label
        ws.cell(row=pr, column=2, value=ref)
        ws.cell(row=pr, column=2).font = font_kpi_sm
        pr += 1

    ws.sub_weekly_apple_contrib = "F4"
    ws.sub_monthly_apple_contrib = "F5"
    ws.sub_weekly_stripe_contrib = "F6"
    ws.sub_monthly_stripe_contrib = "F7"
    return ws


def build_breakeven(wb, lr, sub_ws):
    ws = wb.create_sheet("BREAK-EVEN CALCULATOR")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [28, 14, 14, 14, 14, 14, 14, 14, 14])

    ws["A1"] = "BREAK-EVEN CALCULATOR"
    ws["A1"].font = font_title
    ws.merge_cells("A1:I1")

    fixed_ref = "'EXPENSE TRACKER'!I15"
    cur_ref = f"ASSUMPTIONS!B{lr['Current paid users (total)']}"
    cur_users = as_formula(cur_ref)

    headers = ["Scenario", "Weekly %", "Monthly %", "Apple %", "Avg contrib/user", "Break-even users", "Gross MRR", "Net MRR", "Net profit at current users"]
    ws.append([""])
    ws.append(headers)
    for c in range(1, 10):
        ws.cell(row=3, column=c).font = font_h2
        ws.cell(row=3, column=c).fill = fill_header

    cw = f"'SUBSCRIPTION ECONOMICS'!{sub_ws.sub_weekly_apple_contrib}"
    cm = f"'SUBSCRIPTION ECONOMICS'!{sub_ws.sub_monthly_apple_contrib}"
    cws = f"'SUBSCRIPTION ECONOMICS'!{sub_ws.sub_weekly_stripe_contrib}"
    cms = f"'SUBSCRIPTION ECONOMICS'!{sub_ws.sub_monthly_stripe_contrib}"
    apple_share = f"ASSUMPTIONS!B{lr['Share of subs on Apple (current mix)']}"

    scenarios = [
        ("100% Weekly", 1, 0),
        ("100% Monthly", 0, 1),
        ("75% Weekly / 25% Monthly", 0.75, 0.25),
        ("50% / 50%", 0.5, 0.5),
        ("25% Weekly / 75% Monthly", 0.25, 0.75),
        ("Actual current mix", f"=IF((ASSUMPTIONS!B{lr['Current Weekly subscribers']}+ASSUMPTIONS!B{lr['Current Monthly subscribers']})=0,0.5,ASSUMPTIONS!B{lr['Current Weekly subscribers']}/(ASSUMPTIONS!B{lr['Current Weekly subscribers']}+ASSUMPTIONS!B{lr['Current Monthly subscribers']}))", f"=IF((ASSUMPTIONS!B{lr['Current Weekly subscribers']}+ASSUMPTIONS!B{lr['Current Monthly subscribers']})=0,0.5,ASSUMPTIONS!B{lr['Current Monthly subscribers']}/(ASSUMPTIONS!B{lr['Current Weekly subscribers']}+ASSUMPTIONS!B{lr['Current Monthly subscribers']}))"),
    ]

    r = 4
    for name, wp, mp in scenarios:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=wp)
        ws.cell(row=r, column=3, value=mp)
        ws.cell(row=r, column=4, value=apple_share)
        # Blended contribution: weekly apple * apple% + weekly stripe * (1-apple%) weighted by plan mix
        ws.cell(row=r, column=5, value=f"=B{r}*({apple_share}*{cw}+(1-{apple_share})*{cws})+C{r}*({apple_share}*{cm}+(1-{apple_share})*{cms})")
        ws.cell(row=r, column=6, value=f"=IF(E{r}>0,ROUNDUP({fixed_ref}/E{r},0),\"N/A\")")
        ws.cell(row=r, column=7, value=f"=IF({cur_ref}=0,0,{cur_ref}*(B{r}*ASSUMPTIONS!B{lr['Weekly subscription price']}+C{r}*ASSUMPTIONS!B{lr['Monthly subscription price']}))")
        ws.cell(row=r, column=8, value=f"={cur_ref}*E{r}")
        ws.cell(row=r, column=9, value=f"=H{r}-{fixed_ref}")
        for col in range(5, 10):
            mark_calc(ws.cell(row=r, column=col), "$0.00" if col != 6 else "0")
        r += 1

    ws.cell(row=r + 1, column=1, value="CURRENT PAID USERS").font = font_h2
    ws.cell(row=r + 1, column=2, value=cur_users)
    mark_calc(ws.cell(row=r + 1, column=2), "0")

    ws.cell(row=r + 2, column=1, value="BREAK-EVEN USERS (actual mix scenario)").font = font_h2
    ws.cell(row=r + 2, column=2, value=f"=F{r-1}")
    mark_calc(ws.cell(row=r + 2, column=2), "0")

    ws.cell(row=r + 3, column=1, value="ADDITIONAL USERS STILL NEEDED").font = font_h2
    ws.cell(row=r + 3, column=2, value=f"=MAX(0,B{r+2}-B{r+1})")
    mark_calc(ws.cell(row=r + 3, column=2), "0")
    ws.cell(row=r + 3, column=2).font = font_kpi

    ws.be_row = r + 2
    ws.need_row = r + 3
    ws.actual_profit_row = r - 1
    return ws


def build_recovery(wb, lr, be_ws):
    ws = wb.create_sheet("INVESTMENT RECOVERY")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [36, 18, 36])

    ws["A1"] = "OPERATING BREAK-EVEN vs FULL INVESTMENT RECOVERY"
    ws["A1"].font = font_title
    ws.merge_cells("A1:C1")

    inv = "='TOTAL INVESTMENT'!B9"
    fixed = "='EXPENSE TRACKER'!I15"
    monthly_profit_ref = f"'BREAK-EVEN CALCULATOR'!I{be_ws.actual_profit_row}"
    monthly_profit = as_formula(monthly_profit_ref)

    items = [
        ("A. OPERATING BREAK-EVEN", "", "Monthly net revenue covers fixed + variable at current users"),
        ("Break-even paying users (actual mix)", f"='BREAK-EVEN CALCULATOR'!B{be_ws.be_row}", ""),
        ("Current paid users", f"=ASSUMPTIONS!B{lr['Current paid users (total)']}", ""),
        ("Users still needed", f"='BREAK-EVEN CALCULATOR'!B{be_ws.need_row}", ""),
        ("Current monthly profit / loss (operating)", monthly_profit, "After platform fees + AI at current users"),
        ("Operating status", f"=IF({monthly_profit_ref}>0,\"Profitable\",\"Loss\")", ""),
        ("", "", ""),
        ("B. FULL INVESTMENT RECOVERY", "", "Earn back ALL historical cash invested"),
        ("Total investment to date", inv, ""),
        ("Cumulative net profit (enter monthly in MONTHLY TRACKING)", "=SUM('MONTHLY TRACKING'!Q:Q)", "Sum of monthly profit/loss rows"),
        ("Unrecovered investment", f"=MAX(0,B10-B11)", ""),
        ("Monthly profit after operating break-even", f"=MAX(0,{monthly_profit_ref})", "Uses current month estimate"),
        ("Est. months to recover investment", f"=IF(B13>0,ROUNDUP(B12/B13,0),\"Not yet profitable\")", ""),
    ]

    row = 3
    for label, formula, note in items:
        ws.cell(row=row, column=1, value=label).font = font_h2 if label.startswith(("A.", "B.")) else font_label
        if formula:
            mark_calc(ws.cell(row=row, column=2, value=formula), "$#,##0.00" if "status" not in label.lower() and "months" not in label.lower() else "General")
        ws.cell(row=row, column=3, value=note).font = font_note
        row += 1

    return ws


def build_marketing(wb, lr, sub_ws):
    ws = wb.create_sheet("MARKETING & CAC")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [30, 14, 14, 14, 14, 14, 20])

    ws["A1"] = "MARKETING & CAC CALCULATOR"
    ws["A1"].font = font_title
    ws.merge_cells("A1:G1")

    # Inputs
    ws["A3"] = "Campaign inputs (edit)"
    ws["A3"].font = font_h2
    inputs = [
        ("Marketing budget", 500, "$#,##0.00"),
        ("New weekly subs from campaign", 5, "0"),
        ("New monthly subs from campaign", 2, "0"),
        ("Apple mix for acquired users", f"=ASSUMPTIONS!B{lr['Share of subs on Apple (current mix)']}", "0%"),
    ]
    r = 4
    for label, val, fmt in inputs:
        ws.cell(row=r, column=1, value=label).font = font_label
        mark_input(ws.cell(row=r, column=2, value=val), fmt)
        r += 1

    cw = f"'SUBSCRIPTION ECONOMICS'!{sub_ws.sub_weekly_apple_contrib}"
    cm = f"'SUBSCRIPTION ECONOMICS'!{sub_ws.sub_monthly_apple_contrib}"
    cws = f"'SUBSCRIPTION ECONOMICS'!{sub_ws.sub_weekly_stripe_contrib}"
    cms = f"'SUBSCRIPTION ECONOMICS'!{sub_ws.sub_monthly_stripe_contrib}"
    life = f"ASSUMPTIONS!B{lr['Avg customer lifetime (months)']}"
    apple_share = "B7"

    r += 1
    ws.cell(row=r, column=1, value="CAC (cost per acquired paid user)").font = font_label
    ws.cell(row=r, column=2, value="=IF((B5+B6)>0,B4/(B5+B6),0)")
    mark_calc(ws.cell(row=r, column=2), "$0.00")
    r += 1
    ws.cell(row=r, column=1, value="Blended contribution / user / month").font = font_label
    ws.cell(row=r, column=2, value=f"=IF((B5+B6)=0,0,(B5*({apple_share}*{cw}+(1-{apple_share})*{cws})+B6*({apple_share}*{cm}+(1-{apple_share})*{cms}))/(B5+B6))")
    mark_calc(ws.cell(row=r, column=2), "$0.00")
    r += 1
    ws.cell(row=r, column=1, value="LTV (contribution × lifetime months)").font = font_label
    ws.cell(row=r, column=2, value=f"=B{r-1}*{life}")
    mark_calc(ws.cell(row=r, column=2), "$0.00")
    r += 1
    ws.cell(row=r, column=1, value="LTV : CAC").font = font_label
    ws.cell(row=r, column=2, value="=IF(B8>0,B10/B8,0)")
    mark_calc(ws.cell(row=r, column=2), "0.0")
    r += 1
    ws.cell(row=r, column=1, value="Max affordable CAC (LTV/3 rule)").font = font_label
    ws.cell(row=r, column=2, value="=B10/3")
    mark_calc(ws.cell(row=r, column=2), "$0.00")
    r += 1
    ws.cell(row=r, column=1, value="Payback period (months)").font = font_label
    ws.cell(row=r, column=2, value="=IF(B9>0,B8/B9,0)")
    mark_calc(ws.cell(row=r, column=2), "0.0")
    r += 1
    ws.cell(row=r, column=1, value="Break-even ROAS (need gross rev ≥ spend)").font = font_label
    ws.cell(row=r, column=2, value="=IF(B4>0,(B5*ASSUMPTIONS!B{0}+B6*ASSUMPTIONS!B{1})/B4,0)".format(lr['Weekly subscription price'], lr['Monthly subscription price']))
    mark_calc(ws.cell(row=r, column=2), "0.0%")

    # Ad spend scenarios
    sr = r + 3
    section_title(ws, sr, 1, "AD SPEND SCENARIOS — subs needed to recover campaign", 7)
    sr += 1
    ws.append([])
    headers = ["Ad spend", "Weekly subs needed", "Monthly subs needed", "Mixed (50/50) subs", "Blended contrib", "Payback months @ mixed"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=sr, column=j, value=h).font = font_h2
    sr += 1

    blended_contrib_expr = f"0.5*({apple_share}*{cw}+(1-{apple_share})*{cws})+0.5*({apple_share}*{cm}+(1-{apple_share})*{cms})"
    for spend in (100, 250, 500, 1000):
        ws.cell(row=sr, column=1, value=spend)
        ws.cell(row=sr, column=2, value=f"=IF({cw}>0,ROUNDUP({spend}/{cw},0),\"N/A\")")
        ws.cell(row=sr, column=3, value=f"=IF({cm}>0,ROUNDUP({spend}/{cm},0),\"N/A\")")
        ws.cell(row=sr, column=4, value=f"=IF({blended_contrib_expr}>0,ROUNDUP({spend}/({blended_contrib_expr}),0),\"N/A\")")
        ws.cell(row=sr, column=5, value=f"={blended_contrib_expr}")
        ws.cell(row=sr, column=6, value=f"=IF(E{sr}>0,ROUNDUP({spend}/(E{sr}*ASSUMPTIONS!B{lr['Avg customer lifetime (months)']}),0),\"N/A\")")
        sr += 1

    return ws


def build_monthly_tracking(wb):
    ws = wb.create_sheet("MONTHLY TRACKING")
    ws.sheet_view.showGridLines = False
    headers = [
        "Month", "Weekly Subs", "Monthly Subs", "Apple Subs", "Stripe Subs",
        "New Paid", "Churned", "Active Paid", "Gross Revenue", "Apple Comm",
        "Stripe Fees", "Net Revenue", "Gemini $", "Music Gen $", "Other API $",
        "Total API $", "Infrastructure $", "Marketing $", "Other $", "Total Expense",
        "Profit/Loss", "MRR", "ARPU", "CAC", "LTV", "Distance from BE",
    ]
    set_col_widths(ws, [12] + [11] * (len(headers) - 1))

    ws["A1"] = "MONTHLY TRACKING — add one row per month"
    ws["A1"].font = font_title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    for j, h in enumerate(headers, 1):
        ws.cell(row=3, column=j, value=h).font = font_h2
        ws.cell(row=3, column=j).fill = fill_header

    # Seed 3 historical months + current
    months = ["2025-11", "2025-12", "2026-01", "2026-02"]
    r = 4
    for m in months:
        ws.cell(row=r, column=1, value=m)
        mark_input(ws.cell(row=r, column=1))
        # Formulas for derived cols when user fills subs
        ws.cell(row=r, column=8, value=f"=B{r}+C{r}")  # active ~ sum plan types
        ws.cell(row=r, column=9, value=f"=B{r}*ASSUMPTIONS!B5+C{r}*ASSUMPTIONS!B6")  # rough gross
        ws.cell(row=r, column=16, value=f"=M{r}+N{r}+O{r}")
        ws.cell(row=r, column=20, value=f"=P{r}+Q{r}+R{r}+S{r}")
        ws.cell(row=r, column=21, value=f"=L{r}-T{r}")
        ws.cell(row=r, column=22, value=f"=I{r}")
        ws.cell(row=r, column=23, value=f"=IF(H{r}>0,L{r}/H{r},0)")
        ws.cell(row=r, column=26, value=f"=H{r}-'BREAK-EVEN CALCULATOR'!B12")
        r += 1

    ws.freeze_panes = "A4"
    return ws


def build_dashboard(wb, lr, be_ws, ai_ws):
    ws = wb.create_sheet("FOUNDER DASHBOARD")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [22, 16, 4, 22, 16, 4, 22, 16])

    ws["A1"] = "NABAD AI"
    ws["A1"].font = Font(name="Calibri", size=28, bold=True, color=C_PURPLE)
    ws["A2"] = "Founder Financial Dashboard"
    ws["A2"].font = Font(name="Calibri", size=14, color=C_TEAL)
    ws.merge_cells("A1:H1")
    ws.merge_cells("A2:H2")
    ws["A1"].fill = fill_bg
    ws["A2"].fill = fill_bg

    inv = "='TOTAL INVESTMENT'!B9"
    fixed = "='EXPENSE TRACKER'!I15"
    cur_ref = f"ASSUMPTIONS!B{lr['Current paid users (total)']}"
    cur = as_formula(cur_ref)
    w_sub_ref = f"ASSUMPTIONS!B{lr['Current Weekly subscribers']}"
    m_sub_ref = f"ASSUMPTIONS!B{lr['Current Monthly subscribers']}"
    w_price_ref = f"ASSUMPTIONS!B{lr['Weekly subscription price']}"
    m_price_ref = f"ASSUMPTIONS!B{lr['Monthly subscription price']}"
    be = f"='BREAK-EVEN CALCULATOR'!B{be_ws.be_row}"
    need = f"='BREAK-EVEN CALCULATOR'!B{be_ws.need_row}"
    mprofit = f"='BREAK-EVEN CALCULATOR'!I{be_ws.actual_profit_row}"
    mprofit_ref = embed(mprofit)
    mrr = (
        f"=IF({cur_ref}=0,0,"
        f"{w_sub_ref}*{w_price_ref}+{m_sub_ref}*{m_price_ref})"
    )
    api_cost = (
        f"={w_sub_ref}*'AI UNIT ECONOMICS'!B{ai_ws.meta_avg_weekly_row}+"
        f"{m_sub_ref}*'AI UNIT ECONOMICS'!B{ai_ws.meta_avg_monthly_row}"
    )
    marketing = f"=ASSUMPTIONS!B{lr['Marketing spend this month']}"
    unrec = "='INVESTMENT RECOVERY'!B12"

    kpis = [
        (4, 1, "TOTAL INVESTMENT", inv, "$#,##0.00"),
        (4, 4, "MONTHLY BURN (fixed)", fixed, "$#,##0.00"),
        (4, 7, "MONTHLY API (est.)", api_cost, "$#,##0.00"),
        (7, 1, "CURRENT MRR", mrr, "$#,##0.00"),
        (7, 4, "NET PROFIT / LOSS", mprofit, "$#,##0.00"),
        (7, 7, "MARKETING SPEND", marketing, "$#,##0.00"),
        (10, 1, "PAID USERS", cur, "0"),
        (10, 4, "BREAK-EVEN USERS", be, "0"),
        (10, 7, "STILL NEEDED", need, "0"),
        (13, 1, "WEEKLY $3.99 — Apple contrib", "='SUBSCRIPTION ECONOMICS'!F4", "$0.00"),
        (13, 4, "MONTHLY $9.99 — Apple contrib", "='SUBSCRIPTION ECONOMICS'!F5", "$0.00"),
        (13, 7, "Best Weekly channel", "='SUBSCRIPTION ECONOMICS'!B12", "General"),
        (16, 1, "LTV : CAC", "='MARKETING & CAC'!B11", "0.0"),
        (16, 4, "Max affordable CAC", "='MARKETING & CAC'!B12", "$0.00"),
        (16, 7, "Months to recover investment", "='INVESTMENT RECOVERY'!B14", "General"),
        (19, 1, "OPERATING STATUS", f"=IF({mprofit_ref}>0,\"Profitable\",\"Loss\")", "General"),
        (19, 4, "INVESTMENT STATUS", f"=IF({embed(unrec)}<=0,\"Recovered\",\"Unrecovered \"&TEXT({embed(unrec)},\"$#,##0\"))", "General"),
    ]

    for row, col, label, formula, fmt in kpis:
        kpi_block(ws, row, col, label, formula, fmt)

    # Quick answers
    qr = 22
    section_title(ws, qr, 1, "QUICK ANSWERS", 8)
    qr += 1
    answers = [
        ("Is Weekly $3.99 sustainable?", "='SUBSCRIPTION ECONOMICS'!B10"),
        ("Is Monthly $9.99 sustainable?", "='SUBSCRIPTION ECONOMICS'!B11"),
        ("Apple or Stripe better for Weekly?", "='SUBSCRIPTION ECONOMICS'!B14"),
        ("Apple or Stripe better for Monthly?", "='SUBSCRIPTION ECONOMICS'!B15"),
        ("Main blocker to profitability?", f"=IF({cur_ref}<{embed(be)},\"Low subscriber count\",IF({embed(fixed)}>{mprofit_ref}+100,\"Fixed costs\",\"AI or platform fees\"))"),
        ("Push Weekly or Monthly in marketing?", f"=IF('SUBSCRIPTION ECONOMICS'!F5/'SUBSCRIPTION ECONOMICS'!F4>(ASSUMPTIONS!B{lr['Monthly subscription price']}/ASSUMPTIONS!B{lr['Weekly subscription price']}),\"Monthly (higher contribution)\",\"Weekly (faster trial)\")"),
    ]
    for label, formula in answers:
        ws.cell(row=qr, column=1, value=label).font = font_label
        ws.cell(row=qr, column=2, value=formula).font = font_value
        ws.merge_cells(start_row=qr, start_column=2, end_row=qr, end_column=6)
        qr += 1

    # Chart
    chart = BarChart()
    chart.title = "Contribution margin by plan (Apple)"
    chart.y_axis.title = "USD"
    data = Reference(wb["SUBSCRIPTION ECONOMICS"], min_col=6, min_row=3, max_row=7)
    cats = Reference(wb["SUBSCRIPTION ECONOMICS"], min_col=1, min_row=4, max_row=7)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 16
    chart.height = 9
    ws.add_chart(chart, "A26")

    # Dark styling whole sheet
    for row in ws.iter_rows(min_row=1, max_row=35, min_col=1, max_col=8):
        for cell in row:
            if cell.fill.fill_type is None:
                cell.fill = fill_bg

    ws.sheet_properties.tabColor = C_PURPLE
    return ws


def main():
    wb = Workbook()
    # Remove default sheet later
    ass_ws, lr = build_assumptions(wb)
    exp_ws, exp_r0, exp_tr = build_expenses(wb, lr)
    inv_ws = build_investment(wb, exp_r0, exp_tr)
    ai_ws = build_ai_unit_economics(wb, lr)
    sub_ws = build_subscription_economics(wb, lr, ai_ws)
    be_ws = build_breakeven(wb, lr, sub_ws)
    rec_ws = build_recovery(wb, lr, be_ws)
    mkt_ws = build_marketing(wb, lr, sub_ws)
    mon_ws = build_monthly_tracking(wb)
    dash_ws = build_dashboard(wb, lr, be_ws, ai_ws)

    # Remove default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Reorder: Dashboard first
    wb._sheets.sort(key=lambda s: {
        "FOUNDER DASHBOARD": 0,
        "ASSUMPTIONS": 1,
        "EXPENSE TRACKER": 2,
        "TOTAL INVESTMENT": 3,
        "AI UNIT ECONOMICS": 4,
        "SUBSCRIPTION ECONOMICS": 5,
        "BREAK-EVEN CALCULATOR": 6,
        "INVESTMENT RECOVERY": 7,
        "MARKETING & CAC": 8,
        "MONTHLY TRACKING": 9,
    }.get(s.title, 99))

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
